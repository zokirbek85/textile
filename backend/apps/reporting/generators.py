"""Report generation helpers — Excel (openpyxl) and PDF (reportlab)."""
import io
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL = PatternFill("solid", fgColor="F0F4FA")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _style_data_row(ws, row: int, col_count: int, alternate: bool = False):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        if alternate:
            cell.fill = ALT_FILL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center")


def generate_yarn_cost_excel(batches) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Yarn Cost Report"

    headers = [
        "Batch Code", "Product", "Ne Count", "Type",
        "Start Date", "End Date", "Fiber Input (kg)",
        "Fiber Cost", "Spinning Expenses",
        "Yarn Output (kg)", "Waste (kg)", "Waste %",
        "Cost/kg (UZS)", "Efficiency %",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 25

    for i, batch in enumerate(batches, start=2):
        ws.append([
            batch.batch_code,
            batch.yarn_product.name if batch.yarn_product else "",
            batch.yarn_product.yarn_count if batch.yarn_product else "",
            batch.yarn_product.yarn_type if batch.yarn_product else "",
            batch.start_date.isoformat() if batch.start_date else "",
            batch.end_date.isoformat() if batch.end_date else "",
            float(batch.fiber_input_kg),
            float(batch.fiber_cost_total),
            float(batch.total_spinning_expenses),
            float(batch.yarn_output_kg),
            float(batch.waste_output_kg),
            float(batch.waste_pct),
            float(batch.calculated_yarn_cost_per_kg),
            float(batch.efficiency_pct),
        ])
        _style_data_row(ws, i, len(headers), alternate=i % 2 == 0)

    # Auto-fit column widths
    for col_idx, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_warehouse_balance_excel(ledger_entries) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Warehouse Balances"

    headers = [
        "Warehouse", "Product", "Product Type",
        "Quantity (kg)", "Avg Cost/kg", "Total Value", "Last Movement",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 25

    for i, entry in enumerate(ledger_entries, start=2):
        ws.append([
            entry.warehouse.name,
            entry.product.name,
            entry.product.get_product_type_display(),
            float(entry.quantity_kg),
            float(entry.avg_cost_per_kg),
            float(entry.total_value),
            entry.last_movement_at.date().isoformat() if entry.last_movement_at else "",
        ])
        _style_data_row(ws, i, len(headers), alternate=i % 2 == 0)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_cotton_cost_excel(batches) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotton Fiber Cost Report"

    headers = [
        "Batch Code", "Start Date", "End Date",
        "Cotton Input (kg)", "Cotton Cost",
        "Fiber Output (kg)", "Seed (kg)", "Lint (kg)", "Waste (kg)",
        "Byproduct Credits", "Total Expenses",
        "Net Cost", "Fiber Cost/kg", "Yield %",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for i, batch in enumerate(batches, start=2):
        ws.append([
            batch.batch_code,
            batch.start_date.isoformat() if batch.start_date else "",
            batch.end_date.isoformat() if batch.end_date else "",
            float(batch.cotton_input_kg),
            float(batch.cotton_cost_total),
            float(batch.fiber_output_kg),
            float(batch.seed_output_kg),
            float(batch.lint_output_kg),
            float(batch.waste_output_kg),
            float(batch.total_byproduct_credit),
            float(batch.total_production_expenses),
            float(batch.net_cost),
            float(batch.calculated_fiber_cost_per_kg),
            float(batch.fiber_yield_pct),
        ])
        _style_data_row(ws, i, len(headers), alternate=i % 2 == 0)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
